from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = load_workbook(WORKBOOK_PATH)
sheet_name = 'New Sheet'

worksheet = workbook[sheet_name]

for row in worksheet.iter_rows(min_row=2):
    for column in row:
        print(column.value, end= '\t')
    
        if column.value == 'João':
            worksheet.cell(column.row, 2, 12)
    print()

workbook.save(WORKBOOK_PATH)